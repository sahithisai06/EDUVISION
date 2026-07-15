"""
EduVision — OCR-Based Exam Sheet Reader
Flask backend

Core flow:
  /login        -> simple session auth (gmail-style form)
  /dashboard     -> 4-tile home (Scan Page, My Sheets, Reports, Account)
  /scan          -> opens device camera, captures a frame, sends it here
  /api/extract   -> runs OCR on the captured frame, returns parsed fields for review
  /api/save      -> confirmed record gets appended to the master Excel workbook
  /my-sheets     -> history of everything scanned so far, searchable, exportable
  /account       -> profile + sign out
"""

import os
import re
import io
import base64
import sqlite3
from datetime import datetime

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from PIL import Image
import pytesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\Users\gsail\Downloads\tesseract-ocr-w64-setup-5.5.0.20241111.exe"
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "eduvision.db")
EXCEL_PATH = os.path.join(DATA_DIR, "marks_master.xlsx")

os.makedirs(DATA_DIR, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.environ.get("EDUVISION_SECRET", "dev-secret-change-me")


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            name TEXT,
            created_at TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            roll_number TEXT NOT NULL,
            student_name TEXT,
            department TEXT,
            marks TEXT,
            confidence TEXT,
            scanned_by TEXT,
            created_at TEXT
        )
    """)
    conn.commit()

    existing = conn.execute("SELECT * FROM users WHERE email = ?", ("faculty@cmrcet.ac.in",)).fetchone()
    if not existing:
        conn.execute(
            "INSERT INTO users (email, password_hash, name, created_at) VALUES (?, ?, ?, ?)",
            ("faculty@cmrcet.ac.in", generate_password_hash("eduvision123"), "Demo Faculty",
             datetime.now().isoformat())
        )
        conn.commit()
    conn.close()


def init_excel():
    if os.path.exists(EXCEL_PATH):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks"
    headers = ["Roll Number", "Student Name", "Department", "Marks", "Confidence", "Scanned By", "Timestamp"]
    ws.append(headers)
    header_fill = PatternFill(start_color="3B2F63", end_color="3B2F63", fill_type="solid")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    widths = [16, 24, 14, 10, 12, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(EXCEL_PATH)


init_db()
init_excel()

def login_required(view):
    def wrapped(*args, **kwargs):
        if "user_email" not in session:
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    wrapped.__name__ = view.__name__
    return wrapped


ROLL_PATTERN = re.compile(r"(?:roll\s*(?:no\.?|number)?\s*[:\-]?\s*)([A-Za-z0-9]{6,12})", re.IGNORECASE)
NAME_PATTERN = re.compile(r"(?:name)\s*[:\-]?\s*([A-Za-z .]{2,40})", re.IGNORECASE)
DEPT_PATTERN = re.compile(r"(?:dept\.?|department|branch)\s*[:\-]?\s*([A-Za-z .&]{2,40})", re.IGNORECASE)
MARKS_PATTERN = re.compile(r"(?:marks|score|total)\s*[:\-]?\s*(\d{1,3}(?:\s*/\s*\d{1,3})?)", re.IGNORECASE)
BARE_ROLL_PATTERN = re.compile(r"\b(\d{2}[A-Z]{1}\d{2}[A-Z]{1}\d{2}[A-Z0-9]{2,3})\b", re.IGNORECASE)


def extract_fields(image: Image.Image):
    """Run OCR and pull out roll number / name / department / marks.
    Anything the regex can't find confidently is left blank and flagged
    for the faculty member to fill in manually before saving.
    """
    text = pytesseract.image_to_string(image)

    roll_match = ROLL_PATTERN.search(text) or BARE_ROLL_PATTERN.search(text)
    name_match = NAME_PATTERN.search(text)
    dept_match = DEPT_PATTERN.search(text)
    marks_match = MARKS_PATTERN.search(text)

    fields = {
        "roll_number": roll_match.group(1).upper().strip() if roll_match else "",
        "student_name": name_match.group(1).strip().title() if name_match else "",
        "department": dept_match.group(1).strip().upper() if dept_match else "",
        "marks": marks_match.group(1).strip() if marks_match else "",
        "raw_text": text.strip(),
    }

    found_count = sum(1 for k in ("roll_number", "student_name", "department", "marks") if fields[k])
    if found_count == 4:
        confidence = "High"
    elif found_count >= 2:
        confidence = "Medium — please verify"
    else:
        confidence = "Low — check manually"

    fields["confidence"] = confidence
    return fields


@app.route("/")
def index():
    return redirect(url_for("dashboard") if "user_email" in session else url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
        conn.close()
        if user and check_password_hash(user["password_hash"], password):
            session["user_email"] = user["email"]
            session["user_name"] = user["name"]
            return redirect(url_for("dashboard"))
        return render_template("login.html", error="Incorrect email or password. Try again.")
    if "user_email" in session:
        return redirect(url_for("dashboard"))
    return render_template("login.html", error=None)


@app.route("/signup", methods=["POST"])
def signup():
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "")
    name = request.form.get("name", "").strip() or email.split("@")[0].title()
    if not email or not password:
        return render_template("login.html", error="Email and password are required.")
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO users (email, password_hash, name, created_at) VALUES (?, ?, ?, ?)",
            (email, generate_password_hash(password), name, datetime.now().isoformat())
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return render_template("login.html", error="An account with that email already exists.")
    conn.close()
    session["user_email"] = email
    session["user_name"] = name
    return redirect(url_for("dashboard"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    conn = get_db()
    total = conn.execute("SELECT COUNT(*) c FROM records").fetchone()["c"]
    today = conn.execute(
        "SELECT COUNT(*) c FROM records WHERE date(created_at) = date('now')"
    ).fetchone()["c"]
    conn.close()
    return render_template("dashboard.html", user_name=session.get("user_name"), total=total, today=today)


@app.route("/scan")
@login_required
def scan():
    return render_template("scan.html")


@app.route("/api/extract", methods=["POST"])
@login_required
def api_extract():
    payload = request.get_json(silent=True) or {}
    image_data = payload.get("image", "")
    if not image_data:
        return jsonify({"error": "No image received."}), 400

    try:
        header, encoded = image_data.split(",", 1) if "," in image_data else ("", image_data)
        img_bytes = base64.b64decode(encoded)
        image = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    except Exception:
        return jsonify({"error": "Could not read the captured image."}), 400

    fields = extract_fields(image)
    return jsonify(fields)


@app.route("/api/save", methods=["POST"])
@login_required
def api_save():
    payload = request.get_json(silent=True) or {}
    roll_number = payload.get("roll_number", "").strip()
    student_name = payload.get("student_name", "").strip()
    department = payload.get("department", "").strip()
    marks = payload.get("marks", "").strip()
    confidence = payload.get("confidence", "Manual entry")

    if not roll_number:
        return jsonify({"error": "Roll number is required before saving."}), 400

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    scanned_by = session.get("user_name", session.get("user_email"))

    conn = get_db()
    conn.execute(
        "INSERT INTO records (roll_number, student_name, department, marks, confidence, scanned_by, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (roll_number, student_name, department, marks, confidence, scanned_by, timestamp)
    )
    conn.commit()
    conn.close()

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Marks"]
    ws.append([roll_number, student_name, department, marks, confidence, scanned_by, timestamp])
    wb.save(EXCEL_PATH)

    return jsonify({"status": "saved"})


@app.route("/my-sheets")
@login_required
def my_sheets():
    query = request.args.get("q", "").strip()
    conn = get_db()
    if query:
        rows = conn.execute(
            "SELECT * FROM records WHERE roll_number LIKE ? OR student_name LIKE ? ORDER BY id DESC",
            (f"%{query}%", f"%{query}%")
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM records ORDER BY id DESC").fetchall()
    conn.close()
    return render_template("my_sheets.html", records=rows, query=query)


@app.route("/download-excel")
@login_required
def download_excel():
    return send_file(EXCEL_PATH, as_attachment=True, download_name="EduVision_Marks.xlsx")


@app.route("/account")
@login_required
def account():
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE email = ?", (session["user_email"],)).fetchone()
    scan_count = conn.execute(
        "SELECT COUNT(*) c FROM records WHERE scanned_by = ?", (session.get("user_name"),)
    ).fetchone()["c"]
    conn.close()
    return render_template("account.html", user=user, scan_count=scan_count)


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
